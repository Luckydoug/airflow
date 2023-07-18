import smtplib
from airflow.models import Variable
from calendar import monthrange
import os
from dotenv import load_dotenv
from openpyxl import load_workbook
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import pandas as pd
from colorama import Fore
import random
import calendar

from sub_tasks.libraries.utils import (
    get_month_first_day,
    return_sent_emails,
    record_sent_branch,
    create_initial_file,
    get_todate,
    assert_date_modified,
    clean_folder
)

log = r"/home/opticabi/Documents/optica_reports/targets/log.txt"

load_dotenv()
your_email = os.getenv("douglas_email")
password = os.getenv("douglas_password")
start_date = get_month_first_day()
end_date = get_todate()
receiver_email = ""
scheduled_branches = ["OHO", "YOR"]


date = pd.to_datetime(end_date)
month_number = date.month
year = date.year
days = date.day
number_of_days = monthrange(year, month_number)[1]

fromdate = start_date
dateto = end_date

th_props = [
    ('background-color', '#9ADCFF'),
    ('border', 'solid 1px black'),
    ('font-weight', 'bold'),
    ('text-align', 'center'),
    ('color', 'black'),
    ('padding', '.7em'),
    ('font-size', '13px')
]

tr_props = [
    ('border', 'solid black 1px'),
    ('padding', '.3em')
]

td_props = [
    ('border', 'solid black 1px'),
    ('font-size', '12px'),
    ('font-variant-numeric', 'tabular-nums'),
    ('font-family', 'Georgia, serif')
]

styles = [
    dict(selector="th", props=th_props),
    dict(selector="tr", props=tr_props),
    dict(selector="td", props=td_props)
]

achievement = r"/home/opticabi/airflow/total_branches.xlsx"
achievement = r"/home/opticabi/Documents/optica_reports/total_branches.xlsx"

def send_to_salespersons():
    create_initial_file(filename=log)
    for outlet in scheduled_branches:
        # sales_person = load_workbook(
        #     r"/home/opticabi/airflow/{outlet}.xlsx".format(outlet=outlet)
        #)
        sales_person = load_workbook(
            r"/home/opticabi/Documents/optica_reports/{outlet}.xlsx".format(outlet=outlet))
        payroll_numbers = sales_person.sheetnames
        random_salesperson = random.choice(payroll_numbers)

        # sales = r"/home/opticabi/airflow/{outlet}.xlsx".format(outlet=outlet)
        sales = r"/home/opticabi/Documents/optica_reports/{outlet}.xlsx".format(outlet=outlet)
        files = [sales, achievement]

        branch_achievement = pd.read_excel(achievement, sheet_name=outlet)
        branch_achievement = branch_achievement.style.hide_index().set_properties(
            **{"border-spacing": "0px",  "margin-bottom": "0em", "text-align": "left", "white-space": "normal"}).set_table_styles(styles)
        branch_achievement_html = branch_achievement.to_html(
            doctype_html=True
        )

        if not assert_date_modified(files):
            print(Fore.RED + "The report could not be sent because of the above errors. Kindly clear the errors and try again!")
            print(Fore.WHITE + "Sorry")
            return False
        else:
            for payroll in payroll_numbers:
                if payroll in sales_person.sheetnames:
                    sales_person_report = pd.read_excel(
                        sales, sheet_name=payroll).set_index("Payroll Number")
                    name = sales_person_report.loc[int(payroll), "Name"]
                    print(name)
                    sales_person_email = sales_person_report.loc[int(
                        payroll), "Email"]

                    sales_person_report = sales_person_report[[
                        "Name", "Insurance Target", "Cash Target", "MTD Insurance Sales(count)", "MTD Cash Sales"]]

                    sales_person_report["MTD Achieved Insurance(%)"] = int(sales_person_report.loc[int(
                        payroll), "MTD Insurance Sales(count)"] / ((sales_person_report.loc[int(payroll), "Insurance Target"] * days) / number_of_days) * 100)

                    sales_person_report["MTD Achieved Cash(%)"] = int(round(sales_person_report.loc[int(
                        payroll), "MTD Cash Sales"] / ((sales_person_report.loc[int(payroll), "Cash Target"] * days) / number_of_days) * 100, 0))

                    sales_person_report = sales_person_report[[
                        "Name", "Insurance Target", "MTD Insurance Sales(count)", "MTD Achieved Insurance(%)", "Cash Target", "MTD Cash Sales", "MTD Achieved Cash(%)"]]
                    sales_person_report = sales_person_report.style.hide_index().set_properties(
                        **{"border-spacing": "0px",  "margin-bottom": "0em", "text-align": "left", "white-space": "normal"}).set_table_styles(styles).format({
                            "Cash Target": "{:,d}",
                            "MTD Cash Sales": "{:,d}"
                        })
                    sales_person_report_html = sales_person_report.to_html(
                        doctype_html=True)

                    branch_achievement = pd.read_excel(
                        achievement, sheet_name=outlet)
                    branch_achievement = branch_achievement.style.hide_index().set_properties(
                        **{"border-spacing": "0px",  "margin-bottom": "0em", "text-align": "left", "white-space": "normal"}).set_table_styles(styles)
                    branch_achievement_html = branch_achievement.to_html(
                        doctype_html=True)

                    html = """\
                    <html>
                    <head>
                    <style>
                        table {{border-collapse: collapse;font-family:Times New Roman; font-size:9;}}
                        th {{text-align: left;font-family:Times New Roman; font-size:9; padding: 6px;}}
                        body, p, h3, div, span, var {{font-family:Times New Roman; font-size:11}}
                        td {{text-align: left;font-family:Times New Roman; font-size:9; padding: 8px;}}
                        h4 {{font-size: 12px; font-family: Verdana, sans-serif;}}

                        .content {{
                            margin-top: 20px !important;
                            border: none;
                            background-color: white;
                            padding: 4%;
                            width: 78%;
                            margin: 0 auto;
                            border-radius: 3px;
                        }}

                        .image {{
                            margin-left: -10px;
                            text-align: left;
                            width: 78%;
                            padding: 0
                        }}

                    </style>
                    </head>
                        <body style = "background-color: #F0F0F0; padding: 4% 4% 4% 4%;">
                            <div class = "image">
                            </div>
                            <div class = "content">
                                <h4 style = "color: #00c3ff;">Month to Date Sales</h4>
                                    <div>
                                        <p>Dear <b>{name},</b></p>
                                    </div>
                                    <div>
                                        <p> I hope this finds you well</p>
                                        <p>I hope this message finds you well.
                                        Please find below your sales figures for the month to date. <br> Kindly review and compare them against your sales target.</p>
                                        <table>{sales_person_report_html}</table>
                                        <br>
                                        <p>Also See the overall branch performance</p>
                                        <table style = "width: 80%;">{branch_achievement_html}</table>
                                        <br> <br>
                                        <b>Happy Selling</b> 
                                        <br>
                                        <br>
                                        <b><i>Best Regards</i> <br> <i>{sender_name}</i></b>
                                    </div>
                                    <hr style = "width: 100%; color: gray"/>
                            </div>
                        </body>
                    </html>
                    """.format(
                        name=name.split(" ")[0].capitalize(),
                        branch_achievement_html=branch_achievement_html,
                        sales_person_report_html=sales_person_report_html,
                        sender_name=your_email.split(
                            "@")[0].split('.')[0].capitalize()
                    )

                    if random_salesperson == payroll:
                        receiver_email = [
                            sales_person_email]

                    else:
                        receiver_email = [sales_person_email]

                    # receiver_email = ["tstbranch@gmail.com"]

                    email_message = MIMEMultipart("alternative")
                    email_message["From"] = your_email
                    email_message["Body"] = '<body bgcolor="#fff">text</body>'
                    email_message["To"] = r','.join(receiver_email)
                    email_message["Subject"] = "{name}'s {branch} MTD Performance from {fromdate} to {dateto}".format(
                        fromdate=fromdate, dateto=dateto, branch=outlet, name=name.split(" ")[0].capitalize())
                    email_message.attach(MIMEText(html, "html"))

                    if sales_person_email not in return_sent_emails(filename=log):
                        smtp_server = smtplib.SMTP("smtp.gmail.com", 587)
                        smtp_server.starttls()
                        smtp_server.login(your_email, password)
                        text = email_message.as_string()
                        smtp_server.sendmail(your_email, receiver_email, text)
                        smtp_server.quit()
                        record_sent_branch(
                            sales_person_email,
                            filename=log
                        )
                    else:
                        continue

                else:
                    continue

# send_to_salespersons()

properties = {
    "border-spacing": "0px",
    "margin-bottom": "0em",
    "text-align": "left",
    "white-space": "normal"
}


def send_branch_version():
    for branch in scheduled_branches:
        branch_achievement = pd.read_excel(achievement, sheet_name=branch)
        branch_achievement = branch_achievement.style.hide_index().set_properties(
            **{"border-spacing": "0px",  "margin-bottom": "0em", "text-align": "left", "white-space": "normal"}).set_table_styles(styles)
        branch_achievement_html = branch_achievement.to_html(
            doctype_html=True
        )

        # overall_performance = r"/home/opticabi/airflow/branches.xlsx"
        overall_performance = r"/home/opticabi/Documents/optica_reports/branches.xlsx"
        branch_performance = pd.read_excel(
            overall_performance, index_col=[0, 1], header=[0, 1], sheet_name=branch)
        branch_performance.index = branch_performance.index.get_level_values(0)
        branch_performance = branch_performance.reset_index(level=0)
        branch_performance = branch_performance.rename(
            columns={"Name": "", "Payroll Number": ""})
        branch_performance = branch_performance.rename(
            columns={"": "Name"}, level=1)

        branch_performance = branch_performance.style.hide_index().set_properties(
            **properties).set_table_styles(styles)
        branch_performance_html = branch_performance.to_html(doctype_html=True)

        branch_name = ""
        if not assert_date_modified([overall_performance]):
            return
        else:
            if branch == "OHO":
                receiver_email = [
                    "opticahouse@optica.africa", 
                    "duncan.muchai@optica.africa", 
                    "susan@optica.africa"
                ]
                branch_name = "Optica House"
            elif branch == "YOR":
                receiver_email = [
                    "yorkhouse@optica.africa", 
                    "yh.manager@optica.africa"
                ]
                branch_name = "York House"
            else:
                continue

        html = """
        <!DOCTYPE html>
            <html lang="en">
                <head>
                    <meta charset="UTF-8">
                    <meta http-equiv="X-UA-Compatible" content="IE=edge">
                    <meta name="viewport" content="width=device-width, initial-scale=1.0">
                    <title>Document</title>
                    <style>
                        table {{border-collapse: collapse;font-family:Times New Roman; font-size:6;}}
                        th {{text-align: left;font-family:Times New Roman; font-size:9; padding: 6px;}}
                        body, p, h3, div, span, var {{font-family:Times New Roman; font-size:11}}
                        td {{text-align: left;font-family:Times New Roman; font-size:9; padding: 8px;}}
                        h4 {{font-size: 12px; font-family: Verdana, sans-serif;}}

                        .content {{
                            border: none;
                            background-color: white;
                            padding: 4%;
                            width: 78%;
                            margin: 0 auto;
                            border-radius: 3px;
                        }}
                    </style>
                </head>

                <body style = "background-color: #F0F0F0; padding: 6% 4% 4% 4%;">
                    <div class = "content">
                        <p><b>Hi {branch_name},</b></p>
                        <p>Please see your team performance for this month to Date (MTD)</p>
                        <h4>Overall Branch Performance</h4> <br>
                        <table style = "width: 70%;">{branch_achievement_html}</table>
                        <br>
                        <h4>Individual Performance</h4> <br>
                        <table style = "width: 100%;">{branch_performance_html}</table>
                        <br>
                        <b><i>Best Regards <br> Douglas</i></b>
                    </div>
                </body>

            </html>
        """.format(branch_achievement_html=branch_achievement_html, branch_name=branch_name, branch_performance_html=branch_performance_html)
        email_message = MIMEMultipart("alternative")
        email_message["From"] = your_email
        email_message["To"] = r','.join(receiver_email)
        email_message["Subject"] = f"{branch_name} MTD Performance up to {get_todate().day} {calendar.month_name[get_todate().month].capitalize()} {get_todate().year}"
        email_message.attach(MIMEText(html, "html"))

        smtp_server = smtplib.SMTP("smtp.gmail.com", 587)
        smtp_server.starttls()
        smtp_server.login(your_email, password)
        text = email_message.as_string()
        smtp_server.sendmail(your_email, receiver_email, text)
        smtp_server.quit()



if __name__ == '__main__': 
    send_to_salespersons
    send_branch_version  
    clean_folder
